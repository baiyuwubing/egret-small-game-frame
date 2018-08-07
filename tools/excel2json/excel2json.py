import json
import sys
import os
import time
import openpyxl


def dealValue(value,typeValue):
    if(typeValue=="array"):
        return json.loads(value) 
    else:
        return value

def domain(filename):
    print('domain filename:'+filename)
    basename = os.path.basename(filename)
    fileType=basename.split(".")[-1]
    if(fileType!="xlsx" and fileType!="xls"):
        print('file no .xlsx or xls')
        return

    print 'read file...',time.time()
    workbook = openpyxl.load_workbook(filename) 
    readsheet1 = workbook.active
    print 'read success',time.time()


    keys = []
    keys_type=[]
    for i in xrange(1,readsheet1.max_column+1):
        keys.append(readsheet1.cell(row=2,column=i).value)
        keys_type.append(readsheet1.cell(row=3,column=i).value)

    # print keys
    # print keys_type

    json_data={}
    for i in xrange(4,readsheet1.max_row+1):
        json_data[readsheet1.cell(row=i,column=1).value]={}
        for j in xrange(1,readsheet1.max_column+1):
            value = dealValue(readsheet1.cell(row=i,column=j).value,keys_type[j-1])
            json_data[readsheet1.cell(row=i,column=1).value][keys[j-1]]= value
        
    # print json_data     
    print 'saving...'
    fp = file('./output/'+basename.split(".")[0]+'.json', 'w')
    json.dump(obj=json_data,fp=fp,indent=4)
    print 'save success'
    

if __name__ == '__main__':
    print '======>',time.time()
    # domain('gameConfig.xlsx')
    if(len(sys.argv)>1):
        domain(sys.argv[1])
    else :
        print 'input file name'
    print '<======',time.time()